import pandas as pd
import re
import csv
from datetime import datetime
import matplotlib.pyplot as plt
from collections import defaultdict
from openpyxl.drawing.image import Image
import openpyxl
from openpyxl.styles import Border, Side

def parse_data(regex_time, regex_req, regex_data, path_log, path_for_save):
    log_data  = []
    timestamps = []
    csv_filename = 'tempdata.csv'

    csv_file = open(path_for_save + csv_filename, 'w')

    data = [['time', 'request', 'data']]

    with csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(data)

    csv_file.close()

    with open(path_log) as file:
        for line in file:
            row = []

            try:
                req_time = re.search(regex_time, line).group(0)
                req = re.search(regex_req, line).group(0)
                req_data = re.search(regex_data, line).group(0)

                row.append(req)
                row.append(req_time)
                log_data.append(row)
            except Exception as e:
                continue

            try:
                data = [[req_time, req, req_data]]
                csv_file = open(path_for_save + csv_filename, 'a')
                with csv_file:
                    writer = csv.writer(csv_file)
                    writer.writerows(data)
                    
            except Exception as e:
                continue

    requests_count = defaultdict(lambda: defaultdict(int))

    #Подготовка данных
    requests_count = defaultdict(lambda: defaultdict(int))
    log_data.sort(key=lambda x: datetime.strptime(x[1], "%Y-%m-%d %H:%M:%S"))
    requests, timestamps = zip(*log_data)
    for request, timestamp in zip(requests, timestamps):
        requests_count[timestamp][request] += 1
    unique_requests = list(set(requests))

    #Создание графика
    plt.figure(figsize=(10, 6))
    for i, request in enumerate(unique_requests):
        times = [datetime.strptime(log[1], "%Y-%m-%d %H:%M:%S") 
                for log in log_data if log[0] == request]
        hours = [time.hour + time.minute / 60 for time in times]
        plt.scatter(hours, [request] * len(hours), s=10)

    plt.title("Зависимость времени запроса от типа запроса")
    plt.xlabel("Время (часы)")
    plt.ylabel("Тип запроса")
    plt.xticks(rotation=45)
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(path_for_save + 'scatter_plot.png', bbox_inches='tight', dpi=100)

    #csv -> excel
    read_file = pd.read_csv(path_for_save + csv_filename)
    read_file.to_excel(path_for_save + 'log_file.xlsx', 
                       sheet_name='Данные', 
                       index=False, header=True)

    excel_file = path_for_save + 'log_file.xlsx'
    wb = openpyxl.load_workbook(excel_file)

    sheet = wb['Данные']

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    border_style = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border_style
            
    ws = wb.create_sheet(title="График")

    img = Image(path_for_save + 'scatter_plot.png')
    ws.add_image(img, 'A1')

    wb.save(excel_file)

    plt.show()
